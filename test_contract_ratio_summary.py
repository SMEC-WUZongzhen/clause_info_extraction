import importlib.util
from pathlib import Path

import openpyxl


SCRIPT_PATH = Path(__file__).resolve().parent / "clause_data-flywheel" / "contract_ratio_summary.py"


def load_module():
    spec = importlib.util.spec_from_file_location("contract_ratio_summary", SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def write_json(path: Path, content: str):
    path.write_text(content, encoding="utf-8")


def test_generates_ratio_summary_workbook(tmp_path):
    module = load_module()

    s1_dir = tmp_path / "s1"
    s2_dir = tmp_path / "s2"
    s1_dir.mkdir()
    s2_dir.mkdir()

    contract_a_dir = s1_dir / "合同A"
    contract_a_dir.mkdir()
    write_json(
        contract_a_dir / "chunk1.json",
        """
{
  "task_id": "合同A.md",
  "grouped_result": {
    "安装付款条款": [
      {
        "text": "条款A1",
        "mapped_class": "安装付款条款",
        "context": {"full_context": "上下文A1"}
      }
    ]
  }
}
""".strip(),
    )

    contract_b_dir = s1_dir / "合同B"
    contract_b_dir.mkdir()
    write_json(
        contract_b_dir / "chunk1.json",
        """
{
  "task_id": "合同B.md",
  "grouped_result": {
    "设备付款条款": [
      {
        "text": "条款B1",
        "mapped_class": "设备付款条款",
        "context": {"full_context": "上下文B1"}
      }
    ]
  }
}
""".strip(),
    )

    contract_c_dir = s1_dir / "合同C"
    contract_c_dir.mkdir()
    write_json(
        contract_c_dir / "chunk1.json",
        """
{
  "task_id": "合同C.md",
  "grouped_result": {
    "混签付款条款": [
      {
        "text": "条款C1",
        "mapped_class": "混签付款条款",
        "context": {"full_context": "上下文C1"}
      }
    ]
  }
}
""".strip(),
    )

    write_json(
        s2_dir / "合同A.json",
        """
{
  "results": [
    {
      "extraction_result": [
        {
          "clause_category": "installation_payment",
          "payment_clause": "A-预付款",
          "payment_context": "A上下文1",
          "payment_type": "预付款",
          "payment_ratio": 50.0,
          "payment_amount": null,
          "payment_days": 10
        },
        {
          "clause_category": "installation_payment",
          "payment_clause": "A-尾款",
          "payment_context": "A上下文2",
          "payment_type": "尾款",
          "payment_ratio": 50.0,
          "payment_amount": null,
          "payment_days": 20
        }
      ]
    }
  ]
}
""".strip(),
    )

    write_json(
        s2_dir / "合同B.json",
        """
{
  "results": [
    {
      "extraction_result": [
        {
          "clause_category": "equipment_payment",
          "payment_clause": "B-首款",
          "payment_context": "B上下文1",
          "payment_type": "首款",
          "payment_ratio": 30.0,
          "payment_amount": null,
          "payment_days": 5
        },
        {
          "clause_category": "equipment_payment",
          "payment_clause": "B-金额款",
          "payment_context": "B上下文2",
          "payment_type": "金额款",
          "payment_ratio": null,
          "payment_amount": "5000",
          "payment_days": 15
        }
      ]
    }
  ]
}
""".strip(),
    )

    write_json(
        s2_dir / "合同C.json",
        """
{
  "results": [
    {
      "extraction_result": [
        {
          "clause_category": "equipment_payment",
          "payment_clause": "C-设备首款",
          "payment_context": "C设备上下文1",
          "payment_type": "首款",
          "payment_ratio": 60.0,
          "payment_amount": null,
          "payment_days": 5
        },
        {
          "clause_category": "equipment_payment",
          "payment_clause": "C-设备尾款",
          "payment_context": "C设备上下文2",
          "payment_type": "尾款",
          "payment_ratio": 40.0,
          "payment_amount": null,
          "payment_days": 15
        },
        {
          "clause_category": "installation_payment",
          "payment_clause": "C-安装首款",
          "payment_context": "C安装上下文1",
          "payment_type": "首款",
          "payment_ratio": 30.0,
          "payment_amount": null,
          "payment_days": 8
        },
        {
          "clause_category": "installation_payment",
          "payment_clause": "C-安装尾款",
          "payment_context": "C安装上下文2",
          "payment_type": "尾款",
          "payment_ratio": 40.0,
          "payment_amount": null,
          "payment_days": 18
        }
      ]
    }
  ]
}
""".strip(),
    )

    write_json(
        s2_dir / "合同D.json",
        """
{
  "results": [
    {
      "extraction_result": [
        {
          "clause_category": "equipment_payment",
          "payment_clause": "D-首款",
          "payment_context": "D上下文1",
          "payment_type": "首款",
          "payment_ratio": 70.0,
          "payment_amount": null,
          "payment_days": 8
        },
        {
          "clause_category": "equipment_payment",
          "payment_clause": "D-尾款",
          "payment_context": "D上下文2",
          "payment_type": "尾款",
          "payment_ratio": 40.0,
          "payment_amount": null,
          "payment_days": 18
        }
      ]
    }
  ]
}
""".strip(),
    )

    output_path = tmp_path / "summary.xlsx"
    module.main(["--s1", str(s1_dir), "--s2", str(s2_dir), "--output", str(output_path)])

    workbook = openpyxl.load_workbook(output_path)
    assert workbook.sheetnames == ["S1条款明细", "S2节点明细", "合同统计", "总体汇总"]

    ws_s1 = workbook["S1条款明细"]
    assert [ws_s1.cell(2, col).value for col in range(1, 8)] == [
        "合同A",
        "合同A.md",
        "安装付款条款",
        "条款A1",
        "上下文A1",
        "installation_payment",
        "可统计",
    ]
    mixed_rows = [
        [ws_s1.cell(row, col).value for col in range(1, 8)]
        for row in range(2, ws_s1.max_row + 1)
        if ws_s1.cell(row, 1).value == "合同C"
    ]
    assert mixed_rows == [
        ["合同C", "合同C.md", "混签付款条款", "条款C1", "上下文C1", "equipment_payment", "可统计"],
        ["合同C", "合同C.md", "混签付款条款", "条款C1", "上下文C1", "installation_payment", "可统计"],
    ]

    ws_s2 = workbook["S2节点明细"]
    first_s2_row = [ws_s2.cell(2, col).value for col in range(1, 12)]
    assert first_s2_row == [
        "合同A",
        "installation_payment",
        "A-预付款",
        "A上下文1",
        "预付款",
        50,
        None,
        10,
        100,
        "可统计",
        "达到100%",
    ]

    ws_contract = workbook["合同统计"]
    contract_rows = {
        (ws_contract.cell(row, 1).value, ws_contract.cell(row, 2).value): {
            "ratio_sum": ws_contract.cell(row, 6).value,
            "countable": ws_contract.cell(row, 8).value,
            "result": ws_contract.cell(row, 9).value,
        }
        for row in range(2, ws_contract.max_row + 1)
    }
    assert contract_rows[("合同A", "installation_payment")] == {"ratio_sum": 100, "countable": "是", "result": "达到100%"}
    assert contract_rows[("合同B", "equipment_payment")] == {"ratio_sum": 30, "countable": "否", "result": "不可统计"}
    assert contract_rows[("合同C", "equipment_payment")] == {"ratio_sum": 100, "countable": "是", "result": "达到100%"}
    assert contract_rows[("合同C", "installation_payment")] == {"ratio_sum": 70, "countable": "是", "result": "不足100%"}
    assert contract_rows[("合同D", "equipment_payment")] == {"ratio_sum": 110, "countable": "是", "result": "超过100%"}

    ws_summary = workbook["总体汇总"]
    summary = {ws_summary.cell(row, 1).value: ws_summary.cell(row, 2).value for row in range(2, ws_summary.max_row + 1)}
    assert summary["合同总数"] == 4
    assert summary["统计分组总数"] == 5
    assert summary["可统计分组数目"] == 4
    assert summary["不可统计分组数目"] == 1
    assert summary["达到100%分组数"] == 2
    assert summary["不足100%分组数"] == 1
    assert summary["超过100%分组数"] == 1
    assert summary["达到100%占比"] == "50.00%"
    assert summary["不足100%占比"] == "25.00%"
    assert summary["超过100%占比"] == "25.00%"
