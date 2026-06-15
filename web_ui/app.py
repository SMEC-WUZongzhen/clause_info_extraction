"""app.py — Flask Web 入口
==============================
- GET  /                 首页 (index.html)
- POST /api/upload       上传文件 + 合同类型 → 返回 session_id
- GET  /api/process      SSE 流式推送 step1/step2/error/done
- POST /api/rerun-step2  重新执行 Service 2
- POST /api/export-excel 导出 Step 2 结果到 Excel
"""
from __future__ import annotations

import json
import logging
import queue
import sys
import threading
import time
import uuid
import io
from typing import Dict, Any, List

from flask import Flask, Response, jsonify, render_template, request, send_file
from werkzeug.utils import secure_filename

import config, pipeline
import history_store

# ===== 日志 =====
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("web_ui.app")

# 启动时初始化历史数据库
history_store.init_db()


app = Flask(
    __name__,
    template_folder="templates",
    static_folder="static",
)
app.config["MAX_CONTENT_LENGTH"] = config.MAX_CONTENT_LENGTH

# 会话临时存储：session_id → {"md_bytes": bytes, "contract_type": str, "filename": str, "created_at": float}
SESSIONS: Dict[str, Dict[str, Any]] = {}
_SESSION_LOCK = threading.Lock()
_SESSION_TTL = 300  # 5 分钟

ALLOWED_CONTRACT_TYPES = set(config.PAYMENT_CLASS_MAP.keys())
ALLOWED_OPERATION_TYPES = ("extract", "analyze")

# 前端可配置参数的合法范围
LINES_PER_CHUNK_RANGE = (50, 5000)
MAX_CHARS_RANGE = (50, 5000)


def _safe_error(stage: str, exc: BaseException, public_message: str) -> Dict[str, Any]:
    """记录详细异常到日志，返回脱敏的对外响应体。"""
    logger.exception("[%s] 未预期异常: %s", stage, exc)
    return {"stage": stage, "message": public_message}


def _parse_int_field(raw: str, field_name: str, default: int,
                     valid_range: tuple) -> int:
    """解析并校验整数表单字段；空字符串则用默认值。"""
    if raw is None or str(raw).strip() == "":
        return default
    try:
        v = int(str(raw).strip())
    except ValueError:
        raise ValueError(f"{field_name} 必须为整数")
    lo, hi = valid_range
    if not (lo <= v <= hi):
        raise ValueError(f"{field_name} 需在 [{lo}, {hi}] 范围内")
    return v


def _cleanup_sessions():
    now = time.time()
    with _SESSION_LOCK:
        stale = [sid for sid, s in SESSIONS.items() if now - s["created_at"] > _SESSION_TTL]
        for sid in stale:
            SESSIONS.pop(sid, None)


def _sse(event: str, payload: Any) -> str:
    data = json.dumps(payload, ensure_ascii=False)
    return f"event: {event}\ndata: {data}\n\n"


@app.get("/")
def index():
    return render_template(
        "index.html",
        service2_mode=config.SERVICE2_MODE,
        service2_base_url=config.SERVICE2_CONFIG.get("base_url", ""),
    )


@app.get("/api/runtime-info")
def runtime_info():
    """返回当前 Service 2 连接模式（供前端显示）。"""
    return jsonify({
        "service2_mode": config.SERVICE2_MODE,
        "service2_base_url": config.SERVICE2_CONFIG.get("base_url", ""),
        "service1_base_url": config.SERVICE1_CONFIG.get("base_url", ""),
    })


@app.post("/api/upload")
def upload():
    _cleanup_sessions()

    file = request.files.get("file")
    contract_type = (request.form.get("contract_type") or "").strip()

    if not file or not file.filename:
        return jsonify({"error": "未上传文件"}), 400
    original_filename = file.filename
    if not original_filename.lower().endswith(".md"):
        return jsonify({"error": "仅支持 .md 文件"}), 400
    filename = secure_filename(original_filename)
    # secure_filename 可能剥离纯中文名的 .md 后缀，手动追加而非回退到未净化的原始名
    if not filename.lower().endswith(".md"):
        stem = filename if filename else "uploaded"
        filename = f"{stem}.md"
    if contract_type not in ALLOWED_CONTRACT_TYPES:
        return jsonify({
            "error": f"contract_type 非法（允许: {sorted(ALLOWED_CONTRACT_TYPES)}）"
        }), 400

    # 可选参数：每份行数阈值、上下文最大字符数、是否跳过 Service 2
    try:
        lines_per_chunk = _parse_int_field(
            request.form.get("lines_per_chunk"),
            "lines_per_chunk",
            config.LINES_PER_CHUNK,
            LINES_PER_CHUNK_RANGE,
        )
        max_chars = _parse_int_field(
            request.form.get("max_chars"),
            "max_chars",
            config.MAX_CONTEXT_CHARS,
            MAX_CHARS_RANGE,
        )
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    skip_service2 = (request.form.get("skip_service2") or "").strip().lower() in ("1", "true", "on", "yes")

    # 可选参数：操作模式（extract / analyze）和真值数据
    operation_type = (request.form.get("operation_type") or "extract").strip()
    if operation_type not in ALLOWED_OPERATION_TYPES:
        return jsonify({
            "error": f"operation_type 非法（允许: {list(ALLOWED_OPERATION_TYPES)}）"
        }), 400
    gt_json_raw = (request.form.get("gt_json") or "").strip()
    sis_payment_stages = None
    if operation_type == "analyze" and gt_json_raw:
        try:
            sis_payment_stages = json.loads(gt_json_raw)
            if not isinstance(sis_payment_stages, list):
                return jsonify({"error": "gt_json 必须是 JSON 数组"}), 400
        except json.JSONDecodeError as e:
            return jsonify({"error": f"gt_json 解析失败: {e}"}), 400

    # 可选参数：SIS 合同总金额（仅 analyze 模式有意义；extract 模式忽略）
    sis_contract_price_raw = (request.form.get("sis_contract_price") or "").strip()
    sis_contract_price: float | None = None
    if sis_contract_price_raw:
        try:
            sis_contract_price = float(sis_contract_price_raw)
        except ValueError:
            return jsonify({"error": "sis_contract_price 必须为数字"}), 400

    # 可选参数：任务 ID（留空则在 /api/process 中默认生成 webui-{时间戳}）
    task_id_raw = (request.form.get("task_id") or "").strip()
    if len(task_id_raw) > 128:
        return jsonify({"error": "task_id 长度不能超过 128"}), 400

    md_bytes = file.read()
    if not md_bytes:
        return jsonify({"error": "文件为空"}), 400
    try:
        md_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return jsonify({"error": "文件非 UTF-8 编码"}), 400

    session_id = uuid.uuid4().hex
    with _SESSION_LOCK:
        SESSIONS[session_id] = {
            "md_bytes": md_bytes,
            "contract_type": contract_type,
            "filename": filename,
            "original_filename": original_filename,
            "lines_per_chunk": lines_per_chunk,
            "max_chars": max_chars,
            "skip_service2": skip_service2,
            "operation_type": operation_type,
            "sis_payment_stages": sis_payment_stages,
            "sis_contract_price": sis_contract_price,
            "task_id": task_id_raw,
            "created_at": time.time(),
        }

    return jsonify({
        "session_id": session_id,
        "filename": filename,
        "size": len(md_bytes),
        "contract_type": contract_type,
        "lines_per_chunk": lines_per_chunk,
        "max_chars": max_chars,
        "skip_service2": skip_service2,
        "task_id": task_id_raw,
    })


@app.get("/api/process")
def process():
    _cleanup_sessions()
    session_id = request.args.get("session_id", "").strip()
    with _SESSION_LOCK:
        session = SESSIONS.pop(session_id, None)

    if not session:
        def _err():
            yield _sse("error", {"stage": "init", "message": "session 无效或已过期"})
            yield _sse("done", {})
        return Response(_err(), mimetype="text/event-stream")

    md_bytes = session["md_bytes"]
    contract_type = session["contract_type"]
    filename = session["filename"]
    original_filename = session.get("original_filename", filename)
    lines_per_chunk = session.get("lines_per_chunk") or config.LINES_PER_CHUNK
    max_chars = session.get("max_chars") or config.MAX_CONTEXT_CHARS
    skip_service2 = bool(session.get("skip_service2"))
    operation_type = session.get("operation_type") or "extract"
    sis_payment_stages = session.get("sis_payment_stages")
    sis_contract_price = session.get("sis_contract_price")
    # 优先使用前端指定的 task_id；未指定则默认生成 webui-{时间戳}
    task_id = (session.get("task_id") or "").strip() or f"webui-{int(time.time())}"

    def generate():
        cancelled = threading.Event()
        try:
            yield from _generate_inner(cancelled)
        except GeneratorExit:
            cancelled.set()
            logger.info("SSE client disconnected, cancelling task_id=%s", task_id)
            return

    def _generate_inner(cancelled: threading.Event):
        yield _sse("status", {
            "stage": "step1_running",
            "message": f"Service 1 处理中... (lines={lines_per_chunk}, max_chars={max_chars})",
        })
        # 通过队列在 worker 线程与 SSE 生成器之间转发进度
        progress_q: "queue.Queue[Any]" = queue.Queue()
        SENTINEL = object()
        result_holder: Dict[str, Any] = {}

        def _on_progress(done: int, total: int):
            if cancelled.is_set():
                return
            progress_q.put(("progress", done, total))

        def _worker():
            try:
                result_holder["result"] = pipeline.run_step1(
                    md_bytes, task_id,
                    lines_per_chunk=lines_per_chunk,
                    max_chars=max_chars,
                    on_progress=_on_progress,
                    cancelled=cancelled,
                )
            except BaseException as e:
                result_holder["error"] = e
            finally:
                progress_q.put(SENTINEL)

        worker = threading.Thread(target=_worker, daemon=True)
        worker.start()

        # 实时推送进度事件，直到 worker 完成（通过 SENTINEL 通知）
        # 使用超时轮询以便在客户端断开时及时退出
        while True:
            try:
                msg = progress_q.get(timeout=1.0)
            except queue.Empty:
                if cancelled.is_set():
                    break
                continue
            if msg is SENTINEL:
                break
            _, done, total = msg
            percent = round(done * 100.0 / total, 1) if total else 0.0
            yield _sse("step1_progress", {
                "done": done,
                "total": total,
                "percent": percent,
            })

        worker.join(timeout=5.0)

        err = result_holder.get("error")
        if isinstance(err, pipeline.CancelledError):
            # 客户端断开后，worker 才检测到，正常结束 generator
            return
        if isinstance(err, pipeline.PipelineError):
            yield _sse("error", {"stage": err.stage, "message": err.message, "detail": err.detail})
            yield _sse("done", {})
            return
        if err is not None:
            yield _sse("error", _safe_error("step1", err, "Service 1 处理失败，请查看服务端日志"))
            yield _sse("done", {})
            return

        try:
            step1_result = result_holder["result"]
            paragraphs_raw = step1_result.get("paragraphs", [])
            all_clauses = step1_result.get("all_clauses", [])
            contract_price_info = step1_result.get("contract_price")
            paragraphs = pipeline.apply_contract_type(paragraphs_raw, contract_type)
        except pipeline.PipelineError as e:
            yield _sse("error", {"stage": e.stage, "message": e.message, "detail": e.detail})
            yield _sse("done", {})
            return
        except Exception as e:
            yield _sse("error", _safe_error("step1", e, "Step 1 结果处理失败，请查看服务端日志"))
            yield _sse("done", {})
            return

        yield _sse("step1", {
            "paragraphs": paragraphs,
            "count": len(paragraphs),
            "all_clauses": all_clauses,
            "all_clauses_count": len(all_clauses),
            "contract_type": contract_type,
            "filename": original_filename,
        })

        # 推送合同总价条款（即便为空也推送，便于前端清空旧状态）
        yield _sse("contract_price_clause", {
            "clause": (contract_price_info or {}).get("clause", ""),
            "context": (contract_price_info or {}).get("context", ""),
            "items": (contract_price_info or {}).get("items", []),
            "has_data": bool(contract_price_info),
        })

        if skip_service2:
            yield _sse("status", {"stage": "step2_skipped",
                                   "message": "用户选择仅执行 Service 1，已跳过 Service 2"})
            yield _sse("step2", {"extraction_result": [], "message": "skip: user requested"})
            yield _sse("done", {})
            return

        if not paragraphs:
            yield _sse("status", {"stage": "step2_skipped",
                                   "message": "Step 1 未筛选到付款/质保期条款，跳过 Service 2"})
            yield _sse("step2", {"extraction_result": [], "message": "skip: no paragraphs"})
            yield _sse("done", {})
            return

        if cancelled.is_set():
            return

        yield _sse("status", {"stage": "step2_running", "message": "Service 2 处理中..."})
        try:
            step2 = pipeline.run_step2(
                paragraphs, task_id,
                operation_type=operation_type,
                sis_payment_stages=sis_payment_stages,
            )
        except pipeline.PipelineError as e:
            yield _sse("error", {"stage": e.stage, "message": e.message, "detail": e.detail})
            yield _sse("done", {})
            return
        except Exception as e:
            yield _sse("error", _safe_error("step2", e, "Service 2 处理失败，请查看服务端日志"))
            yield _sse("done", {})
            return

        yield _sse("step2", step2)

        # 合同总金额抽取/比对（独立异常隔离，不阻断 done）
        if contract_price_info and contract_price_info.get("clause"):
            yield _sse("status", {"stage": "contract_price_running",
                                   "message": "Service 2 合同总金额抽取/比对中..."})
            try:
                cp_sis = sis_contract_price if operation_type == "analyze" else None
                cp_result = pipeline.run_compare_contract_price(
                    contract_price_info["clause"],
                    contract_price_info.get("context"),
                    task_id,
                    sis_contract_price=cp_sis,
                )
                yield _sse("contract_price", cp_result)
            except pipeline.PipelineError as e:
                yield _sse("contract_price_error",
                           {"stage": e.stage, "message": e.message, "detail": e.detail})
            except Exception as e:
                body = _safe_error("contract_price", e, "合同总金额比对失败")
                yield _sse("contract_price_error", body)

        yield _sse("done", {})

    headers = {
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",  # 禁用 Nginx 缓冲
    }
    return Response(generate(), mimetype="text/event-stream", headers=headers)


@app.post("/api/rerun-step2")
def rerun_step2():
    """复用前端已缓存的 Step 1 paragraphs，仅重新执行 Service 2。

    Body (JSON):
        {
          "paragraphs": [...],   # 已应用 contract_type 后的 paragraphs（来自前端缓存）
          "task_id": "xxx",      # 可选；留空自动生成 webui-rerun-{时间戳}
          "operation_type": "extract" | "analyze",  # 可选，默认 extract
          "sis_payment_stages": [...]                # analyze 模式需要
        }
    """
    data = request.get_json(silent=True) or {}
    paragraphs = data.get("paragraphs")
    if not isinstance(paragraphs, list):
        return jsonify({"error": "paragraphs 必须为数组"}), 400
    if not paragraphs:
        return jsonify({"error": "paragraphs 为空，无法重跑 Service 2"}), 400

    task_id_raw = (data.get("task_id") or "").strip()
    if len(task_id_raw) > 128:
        return jsonify({"error": "task_id 长度不能超过 128"}), 400
    task_id = task_id_raw or f"webui-rerun-{int(time.time())}"

    operation_type = (data.get("operation_type") or "extract").strip()
    if operation_type not in ALLOWED_OPERATION_TYPES:
        return jsonify({
            "error": f"operation_type 非法（允许: {list(ALLOWED_OPERATION_TYPES)}）"
        }), 400
    sis_payment_stages = data.get("sis_payment_stages") if operation_type == "analyze" else None

    # 可选：合同总价条款（提供时重跑 Service 2 后再调用一次 /compare_contract_price）
    cp_clause = (data.get("contract_price_clause") or "").strip()
    cp_context = data.get("contract_price_clause_context") or None
    cp_sis_raw = data.get("sis_contract_price")
    cp_sis: float | None = None
    if cp_sis_raw is not None and str(cp_sis_raw).strip() != "":
        try:
            cp_sis = float(cp_sis_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "sis_contract_price 必须为数字"}), 400

    try:
        result = pipeline.run_step2(
            paragraphs, task_id,
            operation_type=operation_type,
            sis_payment_stages=sis_payment_stages,
        )
    except pipeline.PipelineError as e:
        return jsonify({"error": e.message, "stage": e.stage, "detail": e.detail}), 502
    except Exception as e:
        body = _safe_error("step2", e, "Service 2 处理失败，请查看服务端日志")
        return jsonify({"error": body["message"], "stage": body["stage"]}), 500

    response = {"task_id": task_id, "result": result}

    if cp_clause:
        try:
            cp_result = pipeline.run_compare_contract_price(
                cp_clause, cp_context, task_id,
                sis_contract_price=cp_sis if operation_type == "analyze" else None,
            )
            response["contract_price"] = cp_result
        except pipeline.PipelineError as e:
            response["contract_price_error"] = {
                "stage": e.stage, "message": e.message, "detail": e.detail,
            }
        except Exception as e:
            body = _safe_error("contract_price", e, "合同总金额比对失败")
            response["contract_price_error"] = body

    return jsonify(response)


@app.post("/api/compare-contract-price")
def api_compare_contract_price():
    """供 analyze 模式「重新对比」按钮调用。

    Body (JSON):
        {
          "clause": "合同总价条款原文（必填）",
          "context": "上下文（可选）",
          "sis_contract_price": 120000,  # 可选，缺省则仅抽取
          "task_id": "xxx"               # 可选
        }
    """
    data = request.get_json(silent=True) or {}
    clause = (data.get("clause") or "").strip()
    if not clause:
        return jsonify({"error": "clause 不能为空"}), 400

    context = data.get("context") or None

    sis_raw = data.get("sis_contract_price")
    sis_value: float | None = None
    if sis_raw is not None and str(sis_raw).strip() != "":
        try:
            sis_value = float(sis_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "sis_contract_price 必须为数字"}), 400

    task_id_raw = (data.get("task_id") or "").strip()
    if len(task_id_raw) > 64:
        return jsonify({"error": "task_id 长度不能超过 64"}), 400
    task_id = task_id_raw or f"webui-cp-{int(time.time())}"

    try:
        result = pipeline.run_compare_contract_price(
            clause, context, task_id, sis_contract_price=sis_value,
        )
    except pipeline.PipelineError as e:
        return jsonify({"error": e.message, "stage": e.stage, "detail": e.detail}), 502
    except Exception as e:
        body = _safe_error("contract_price", e, "合同总金额比对失败，请查看服务端日志")
        return jsonify({"error": body["message"], "stage": body["stage"]}), 500

    return jsonify({"task_id": task_id, "result": result})


# ===== 字段定义（供 Excel 导出） =====
PAYMENT_FIELDS = [
    ("_source_file", "来源文件"),
    ("_source_time", "来源时间"),
    ("_source_type", "合同类型"),
    ("clause_category", "类别"),
    ("payment_type", "阶段类型"),
    ("payment_code", "节点编码"),
    ("payment_ratio", "比例"),
    ("payment_amount", "金额"),
    ("payment_days", "付款天数"),
    ("latest_payment_stage", "最迟付款节点"),
    ("latest_payment_date", "最迟付款时间(天)"),
    ("payment_clause", "条款原文"),
    ("payment_context", "上下文"),
]
WARRANTY_FIELDS = [
    ("_source_file", "来源文件"),
    ("_source_time", "来源时间"),
    ("_source_type", "合同类型"),
    ("warranty", "质保期"),
    ("warranty_clause", "条款原文"),
]


@app.post("/api/export-excel")
def export_excel():
    """导出 Step 2 结果到 Excel 文件。

    Body (JSON):
        {
          "payment_items": [...],
          "warranty_items": [...],
          "payment_fields": ["payment_type", "payment_ratio", ...],
          "warranty_fields": ["warranty", "warranty_clause"],
        }
    """
    import openpyxl

    data = request.get_json(silent=True) or {}
    payment_items = data.get("payment_items") or []
    warranty_items = data.get("warranty_items") or []
    selected_payment_fields = data.get("payment_fields") or []
    selected_warranty_fields = data.get("warranty_fields") or []

    if not payment_items and not warranty_items:
        return jsonify({"error": "无数据可导出"}), 400

    pay_fields = [(k, v) for k, v in PAYMENT_FIELDS if k in selected_payment_fields]
    war_fields = [(k, v) for k, v in WARRANTY_FIELDS if k in selected_warranty_fields]

    wb = _build_export_workbook([
        ("付款条款", payment_items, pay_fields),
        ("质保期", warranty_items, war_fields),
    ])
    if wb is None:
        return jsonify({"error": "无可导出的字段"}), 400

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    filename = f"payment_export_{timestamp}.xlsx"

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _build_export_workbook(sheets):
    """构造 Excel 工作簿。sheets: List[(sheet_name, items, fields)]。

    没有任何可写 sheet 时返回 None。
    """
    import openpyxl

    populated = [(name, items, fields) for name, items, fields in sheets if items and fields]
    if not populated:
        return None

    wb = openpyxl.Workbook()
    # 移除默认 sheet，统一用 create_sheet 显式创建，避免名称/状态二义
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, items, fields in populated:
        ws = wb.create_sheet(sheet_name)
        ws.append([label for _, label in fields])
        for item in items:
            row = [str(item.get(k, "") or "") for k, _ in fields]
            ws.append(row)
        for col_idx, _ in enumerate(fields, 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 60)
    return wb


# ============================================================
#  历史记录 API
# ============================================================

@app.get("/api/history")
def api_history_list():
    """列表（仅摘要，不含大 JSON 字段）。可通过 ?type=extract|compare 过滤。"""
    op_type = request.args.get("type", "").strip() or None
    records = history_store.list_records(operation_type=op_type)
    return jsonify(records)


@app.get("/api/history/<record_id>")
def api_history_get(record_id: str):
    """获取单条完整记录。"""
    record = history_store.get_record(record_id)
    if record is None:
        return jsonify({"error": "记录不存在"}), 404
    return jsonify(record)


@app.post("/api/history")
def api_history_create():
    """创建一条历史记录（SSE done 后前端调用）。"""
    data = request.get_json(silent=True) or {}
    if not data.get("id"):
        return jsonify({"error": "缺少 id 字段"}), 400
    try:
        history_store.create_record(data)
    except history_store.DuplicateRecordError:
        return jsonify({"error": "记录 id 已存在", "id": data["id"]}), 409
    return jsonify({"ok": True, "id": data["id"]})


@app.delete("/api/history/<record_id>")
def api_history_delete(record_id: str):
    """删除单条记录。"""
    deleted = history_store.delete_record(record_id)
    return jsonify({"ok": True, "deleted": deleted})


@app.delete("/api/history")
def api_history_clear():
    """清空记录。可通过 ?type=extract|compare 仅清空指定类型。"""
    op_type = request.args.get("type", "").strip() or None
    count = history_store.delete_all(operation_type=op_type)
    return jsonify({"ok": True, "deleted_count": count})


@app.post("/api/history/export")
def api_history_export():
    """批量导出勾选的历史记录到 Excel。"""
    body = request.get_json(silent=True) or {}
    ids = body.get("ids") or []
    if not ids:
        return jsonify({"error": "请选择要导出的记录"}), 400

    records = history_store.get_records_by_ids(ids)
    if not records:
        return jsonify({"error": "未找到匹配的记录"}), 404

    # 收集数据，为每条添加来源标识
    all_payment_items = []
    all_warranty_items = []
    for rec in records:
        src_file = rec.get("filename", "")
        src_time = rec.get("createdAt", "")
        src_type = rec.get("contractTypeLabel", "")
        for it in rec.get("paymentItems", []):
            all_payment_items.append({
                **it,
                "_source_file": src_file,
                "_source_time": src_time,
                "_source_type": src_type,
            })
        for it in rec.get("warrantyItems", []):
            all_warranty_items.append({
                **it,
                "_source_file": src_file,
                "_source_time": src_time,
                "_source_type": src_type,
            })

    if not all_payment_items and not all_warranty_items:
        return jsonify({"error": "所选记录中无 Step 2 数据"}), 400

    wb = _build_export_workbook([
        ("付款条款", all_payment_items, list(PAYMENT_FIELDS)),
        ("质保期", all_warranty_items, list(WARRANTY_FIELDS)),
    ])
    if wb is None:
        return jsonify({"error": "无可导出的字段"}), 400

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"history_export_{timestamp}.xlsx",
    )


@app.errorhandler(413)
def too_large(_e):
    return jsonify({"error": f"文件过大（> {config.MAX_CONTENT_LENGTH} 字节）"}), 413


def _parse_cli():
    import argparse
    parser = argparse.ArgumentParser(
        description="合同付款信息提取 - 可视化前端 Web 服务",
    )
    parser.add_argument(
        "--mode", "--service2-mode",
        dest="service2_mode",
        choices=["local", "remote"],
        default= "remote",
        help="Service 2 连接模式：local(localhost:8001) / remote(百舸平台)；"
             "默认读取环境变量 SERVICE2_MODE，未设置则为 remote",
    )
    parser.add_argument("--host", default=None, help=f"监听地址（默认 {config.HOST}）")
    parser.add_argument("--port", type=int, default=None, help=f"监听端口（默认 {config.PORT}）")
    return parser.parse_args()


if __name__ == "__main__":
    _args = _parse_cli()

    if _args.service2_mode:
        active = config.set_service2_mode(_args.service2_mode)
    else:
        active = config.SERVICE2_CONFIG

    # 安全：remote 模式必须显式注入 API Key
    if config.SERVICE2_MODE == "remote" and not active.get("api_key"):
        logger.error(
            "SERVICE2_MODE=remote 但未设置 SERVICE2_API_KEY 环境变量；"
            "请通过环境变量注入后重新启动。"
        )
        sys.exit(2)

    host = _args.host or config.HOST
    # 端口默认按 --mode 区分：remote=5000，local=5001；用户显式 --port 则覆盖
    if _args.port is not None:
        port = _args.port
    else:
        port = 5000 if config.SERVICE2_MODE == "remote" else 5001

    logger.info("=" * 60)
    logger.info("Service 1 URL : %s", config.SERVICE1_CONFIG['base_url'])
    logger.info("Service 2 MODE: %s", config.SERVICE2_MODE)
    logger.info("Service 2 URL : %s%s", active['base_url'], active['endpoint'])
    logger.info("Web listening : http://%s:%s", host, port)
    logger.info("=" * 60)

    app.run(host=host, port=port, debug=False, threaded=True)
