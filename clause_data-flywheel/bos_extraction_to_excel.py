"""
直接从 BOS 读取每个子目录下的 extraction_result.json，
按日期过滤后将所有条款汇总到一个 Excel 文件，不在本地保留 JSON。
"""
import json
import os
import pandas as pd
from datetime import datetime, timezone
from baidubce.services.bos.bos_client import BosClient
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import bos_sample_conf

bos_client = BosClient(bos_sample_conf.config)

bucket_name = r'smec-attachment-ai-supervise'
file_path = r'smec-attachment-ai-supervise/contract-risk/payment-extraction'
output_excel = r'E:\DEMO_CODE\付款条款节点提取服务_smec\clause_data-flywheel\extraction_summary.xlsx'

# 只处理此日期（UTC）之后修改的文件，设为 None 则不过滤
after_date = datetime(2026, 6, 24, tzinfo=timezone.utc)

# 只处理文件名为该值的对象
target_filename = 'extraction_result.json'

columns = [
    'id',
    'file_created_at',
    'clause_category',
    'payment_clause',
    'payment_context',
    'payment_type',
    'payment_code',
    'payment_ratio',
    'payment_amount',
    'payment_days',
    'latest_payment_stage',
    'latest_payment_date',
    'special_clause_content',
]

# BOS 对象 key 不含 bucket_name 前缀，需去掉
prefix = file_path
if prefix.startswith(bucket_name + '/'):
    prefix = prefix[len(bucket_name) + 1:]

rows = []
total, skipped_date, skipped_err = 0, 0, 0
marker = None

while True:
    kwargs = {'prefix': prefix}
    if marker:
        kwargs['marker'] = marker

    response = bos_client.list_objects(bucket_name, **kwargs)

    for obj in response.contents:
        key = obj.key

        # 只处理目标文件名
        if os.path.basename(key) != target_filename:
            continue

        # 日期过滤
        if after_date is not None:
            last_modified = datetime.strptime(obj.last_modified, '%Y-%m-%dT%H:%M:%SZ').replace(tzinfo=timezone.utc)
            if last_modified <= after_date:
                skipped_date += 1
                continue
        else:
            last_modified = datetime.strptime(obj.last_modified, '%Y-%m-%dT%H:%M:%SZ').replace(tzinfo=timezone.utc)

        file_created_at = last_modified.strftime('%Y-%m-%d %H:%M:%S')

        # 直接从 BOS 读 JSON 内容到内存（不落盘）
        try:
            resp = bos_client.get_object(bucket_name, key)
            content = resp.data.read().decode('utf-8')
            data = json.loads(content)
        except Exception as e:
            print(f'[SKIP] {key}: {e}')
            skipped_err += 1
            continue

        record_id = data.get('id')
        extraction_result = data.get('extraction_result') or []

        if not extraction_result:
            rows.append({col: (record_id if col == 'id' else (file_created_at if col == 'file_created_at' else None)) for col in columns})
            total += 1
            continue

        for item in extraction_result:
            row = {col: item.get(col) for col in columns}
            row['id'] = record_id
            row['file_created_at'] = file_created_at
            rows.append(row)
            total += 1

        print(f'[OK] {key} -> {len(extraction_result)} clauses')

    if response.is_truncated:
        marker = response.next_marker
    else:
        break

df = pd.DataFrame(rows, columns=columns)
os.makedirs(os.path.dirname(output_excel), exist_ok=True)
df.to_excel(output_excel, index=False)

# === 按 id 列对相同 id 的行着色 ===
color_list = [
    "FFF2CC", "D9EAD3", "D0E0E3", "F4CCCC", "CFE2F3",
    "EAD1DC", "FCE5CD", "E2EFDA", "F9CB9C", "D9D2E9",
]
wb = load_workbook(output_excel)
ws = wb.active

header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
col_idx = header.index('id')

header_fill = PatternFill("solid", fgColor="BDD7EE")
for cell in ws[1]:
    cell.fill = header_fill

id2color = {}
color_idx = 0
for row in ws.iter_rows(min_row=2):
    id_val = row[col_idx].value
    if id_val not in id2color:
        id2color[id_val] = color_list[color_idx % len(color_list)]
        color_idx += 1
    fill = PatternFill("solid", fgColor=id2color[id_val])
    for cell in row:
        cell.fill = fill

wb.save(output_excel)

print(f'\nDone. Rows: {total}, Skipped(date): {skipped_date}, Skipped(error): {skipped_err}')
print(f'Output: {output_excel}')
